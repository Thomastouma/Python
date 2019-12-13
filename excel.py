import back2
import openpyxl


def write_to_default_ws_row_col():
    excel_data = back2.extract_excel_data()
    wb = openpyxl.load_workbook('thomas löneunderlag.xlsx')
    sheet = wb.get_sheet_by_name('1')

    for rowNum in range(4, sheet.max_row):  # skip the three first rows
        dag = sheet.cell(row=rowNum, column=1).value
        for datum in excel_data:
            if datum[0] == str(dag):   #remember to reformat database column datum to int value, so i can compare easier
                sheet.cell(row=rowNum, column=2).value = datum[1]

    wb.save('updatedthomaslöneunderlag.xlsx')
    #wb.save('thomas löneunderlag.xlsx')
